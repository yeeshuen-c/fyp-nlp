import asyncio
from datetime import datetime
import os
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
import pandas as pd
from bson import ObjectId

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

async def update_user_passwords():
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]
    users = await db.users.find().to_list(length=None)
    for user in users:
        if not user["password"].startswith("$2b$"):  # Check if the password is not hashed
            hashed_password = pwd_context.hash(user["password"])
            await db.users.update_one({"_id": user["_id"]}, {"$set": {"password": hashed_password}})
    print("Passwords updated successfully.")

async def export_post_labelling():
    # Connect to MongoDB
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Fetch posts and users data
    posts = await db.posts.find().to_list(length=None)
    users = await db.users.find().to_list(length=None)

    # Create a dictionary for quick lookup of user details by user_id
    user_dict = {user["user_id"]: user for user in users}

    # Prepare data for Excel
    data = []
    for post in posts:
        user_id = post.get("user_id")
        user = user_dict.get(user_id, {})
        agency_name = user.get("agencies_name", "N/A")
        platform_name = user.get("platform", {}).get("platform_name", "N/A")
        scam_type = post.get("analysis", {}).get("scam_type", "N/A")
        scam_framing = post.get("analysis", {}).get("scam_framing", "N/A")

        data.append({
            "post_id": post.get("post_id", "N/A"),
            "content": post.get("content", "N/A"),
            "user_id": user_id,
            "agency name": agency_name,
            "platform name": platform_name,
            "scam type": scam_type,
            "scam prevention framing": scam_framing
        })

    # Convert to DataFrame
    df = pd.DataFrame(data)

    # Ensure the 'excel' directory exists
    output_dir = "excel"
    os.makedirs(output_dir, exist_ok=True)

    # Export to Excel
    output_file = os.path.join(output_dir, "post_labelling.xlsx")

    df.to_excel(output_file, index=False)
    print(f"Data exported successfully to {output_file}")


def compare_excel_files():
    # File paths
    excel_dir = "excel"
    file1 = os.path.join(excel_dir, "post_labelling.xlsx")
    file2 = os.path.join(excel_dir, "scam_prevention_posts (osb).xlsx")
    output_file = os.path.join(excel_dir, "difference.xlsx")

    # Load the Excel files
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Normalize the columns for comparison (ignore case and strip whitespace)
    df1["scam type"] = df1["scam type"].str.lower().str.strip()
    df1["scam prevention framing"] = df1["scam prevention framing"].str.lower().str.strip()
    df1["content"] = df1["content"].str.lower().str.strip()
    df2["scam type"] = df2["scam type"].str.lower().str.strip()
    df2["scam prevention framing"] = df2["scam prevention framing"].str.lower().str.strip()
    df2["content"] = df2["content"].str.lower().str.strip()

    # Merge the two DataFrames on `content` to align rows
    merged_df = pd.merge(df1, df2, on="content", suffixes=("_file1", "_file2"))

    # Find rows with differences in `scam type` or `scam prevention framing`
    differences = merged_df[
        (merged_df["scam type_file1"] != merged_df["scam type_file2"]) |
        (merged_df["scam prevention framing_file1"] != merged_df["scam prevention framing_file2"])
    ]

    # Select relevant columns for the output
    output_df = differences[[
        "content",
        "scam type_file1", "scam prevention framing_file1",
        "scam type_file2", "scam prevention framing_file2"
    ]]

    # Save the differences to a new Excel file
    output_df.to_excel(output_file, index=False)
    print(f"Differences exported successfully to {output_file}")

async def count_documents_by_user_id():
    """
    Count documents in the database based on user_id ranges and output the counts.
    """
    # Connect to MongoDB
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Define ranges and labels
    ranges = {
        "official website": {"min": 1, "max": 5},
        "twitter": {"min": 6, "max": 10},
        "facebook": {"min": 11, "max": 15},
    }

    # Initialize counts
    counts = {label: 0 for label in ranges}

    # Iterate over ranges and count documents
    for label, range_values in ranges.items():
        count = await db.posts.count_documents({
            "user_id": {"$gte": range_values["min"], "$lte": range_values["max"]}
        })
        counts[label] = count

    # Print the counts
    for label, count in counts.items():
        print(f"{label.capitalize()} count: {count}")

async def update_comment_counts_for_posts():
    """
    Update the engagement.comment_count field in the posts collection
    for user_id between 6 and 10 based on the actual comment count
    from the comments collection.
    """
    # Connect to MongoDB
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Find posts with user_id between 6 and 10
    posts = await db.posts.find({"user_id": {"$gte": 6, "$lte": 10}}).to_list(length=None)

    for post in posts:
        post_id = post.get("_id")
        if not post_id:
            continue

        # Find all comments linked to the current post_id
        comments = await db.comments.find({"post_id": post_id}).to_list(length=None)

        # Calculate the total number of comment_content entries across all comments documents
        total_comment_content_count = sum(len(comment.get("comments", [])) for comment in comments)

        # Update the engagement.comment_count2 field in the posts collection
        await db.posts.update_one(
            {"_id": post_id},
            {"$set": {"engagement.comment_count2": total_comment_content_count}}
        )

        print(f"Updated post_id {post_id} with comment_count {total_comment_content_count}")

    print("Comment counts updated successfully.")

from openpyxl import load_workbook

def update_column_r():
    """
    Read column Q from 'Scam_Coding_Sheet_Comparison_20250409.xlsx' (rows 3 to 1160)
    and replace column R in 'Scam_Coding_Sheet_Comparison2.xlsx' (rows 3 to 1160),
    while preserving original formulas in the Excel file.
    """
    # File paths
    file1 = r"C:\Users\yscha\OneDrive - matrik\USM\Y4S1\FYP\code\bend\excel\Scam_Coding_Sheet_Comparison_20250409.xlsx"
    file2 = r"C:\Users\yscha\OneDrive - matrik\USM\Y4S1\FYP\code\bend\excel\Scam_Coding_Sheet_Comparison2.xlsx"
    output_file = r"C:\Users\yscha\OneDrive - matrik\USM\Y4S1\FYP\code\bend\excel\Scam_Coding_Sheet_Comparison2_Updated.xlsx"

    # Load both Excel files
    wb1 = load_workbook(file1, data_only=False)  # Load source file (keep formulas)
    wb2 = load_workbook(file2, data_only=False)  # Load target file (keep formulas)

    # Select the first sheet in both workbooks
    ws1 = wb1.active
    ws2 = wb2.active

    # Copy column Q (index 17 in Excel, 16 in 0-based indexing) from file1 to column R (index 18 in Excel, 17 in 0-based indexing) in file2
    for row in range(3, 1161):  # Rows 3 to 1160
        ws2.cell(row=row, column=18).value = ws1.cell(row=row, column=17).value

    # Save the updated file2 to a new Excel file
    wb2.save(output_file)
    print(f"Column R updated successfully and saved to {output_file}")

async def update_scam_framing():
    """
    Read the Excel file, decrement post_id by 1, find the corresponding MongoDB document,
    and update the analysis.scam_framing2 field if the content matches.
    """
    # File path
    file = r"C:\Users\yscha\OneDrive - matrik\USM\Y4S1\FYP\code\bend\excel\Scam_Coding_Sheet_Comparison2_Updated.xlsx"

    # Load the Excel file
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # Connect to MongoDB
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Iterate through rows in the Excel file (starting from row 3)
    for row in range(171, ws.max_row + 1):
        post_id = ws.cell(row=row, column=1).value  # Column A (post_id)
        post_text = ws.cell(row=row, column=2).value  # Column B (post_text)
        scam_framing2 = ws.cell(row=row, column=16).value  # Column P (scam_framing2)

        if post_id is None or post_text is None or scam_framing2 is None:
            continue  # Skip rows with missing data

        # # Decrement post_id by 1
        # post_id -= 1

        # Find the document in MongoDB
        document = await db.posts.find_one({"post_id": post_id})

        if document:
            # Update the analysis.scam_framing2 field
            await db.posts.update_one(
                {"post_id": post_id},
                {"$set": {"analysis.scam_framing2": scam_framing2}}
            )
            print(f"Updated post_id {post_id} with scam_framing2: {scam_framing2}")
        else:
            print(f"No document found for post_id {post_id}")

    print("Scam framing updates completed.")    

async def update_sentiment_analysis_by_comment_id():
    """
    Read the Excel file, find MongoDB documents using comment_id from column A,
    check if the comment_content matches the text in column C, and update
    the sentiment_analysis2 field in the comments array with the value from column G.
    """
    # File path
    file = r"C:\Users\yscha\OneDrive - matrik\USM\Y4S1\FYP\code\bend\excel\sentiment_analysis_hf.xlsx"

    # Load the Excel file
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # Connect to MongoDB
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Iterate through rows in the Excel file (starting from row 2 to skip the header)
    for row in range(2, ws.max_row + 1):
        comment_id = ws.cell(row=row, column=1).value  # Column A (comment_id)
        comment_text = ws.cell(row=row, column=3).value  # Column C (text)
        sentiment_analysis2 = ws.cell(row=row, column=7).value  # Column G (sentiment_analysis2)

        if comment_id is None or comment_text is None or sentiment_analysis2 is None:
            continue  # Skip rows with missing data

        # Find the document in MongoDB using the comment_id
        document = await db.comments.find_one({"comment_id": comment_id})

        if document:
            # Iterate through the comments array in the document
            comments = document.get("comments", [])
            for comment in comments:
                if comment.get("comment_content") == comment_text:
                    # Update the sentiment_analysis2 field for the matching comment
                    await db.comments.update_one(
                        {"comment_id": comment_id, "comments.comment_content": comment_text},
                        {"$set": {"comments.$.sentiment_analysis2": sentiment_analysis2}}
                    )
                    print(f"Updated comment_id {comment_id} with sentiment_analysis2: {sentiment_analysis2}")
                    break
            else:
                print(f"No matching comment_content found for comment_id {comment_id}")
        else:
            print(f"No document found for comment_id {comment_id}")

    print("Sentiment analysis updates completed.")

async def fix_scam_type_none():
    """
    Find post documents where analysis.scam_type is "Non", "None ", or "none"
    and update it to "None".
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Find all posts with scam_type needing correction
    query = {
        "$or": [
            {"analysis.scam_type": "Non"},
            {"analysis.scam_type": "None "},
            {"analysis.scam_type": "none"}
        ]
    }
    posts = await db.posts.find(query).to_list(length=None)
    print(f"Found {len(posts)} posts to update.")

    for post in posts:
        post_id = post.get("post_id")
        await db.posts.update_one(
            {"_id": post["_id"]},
            {"$set": {"analysis.scam_type": "None"}}
        )
        print(f"Updated post_id {post_id} scam_type to 'None'.")

    print("Scam type normalization completed.")

async def fix_scam_framing():
    """
    Find post documents where analysis.scam_type is "-"
    and update it to "None".
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Find all posts with scam_type needing correction
    query = {
        "$or": [
            {"analysis.scam_framing2": "-"},
            {"analysis.scam_framing2": "none"},
            {"analysis.scam_framing2": "None"},
            {"analysis.scam_framing2": "Non"}
        ]
    }
    posts = await db.posts.find(query).to_list(length=None)
    print(f"Found {len(posts)} posts to update.")

    for post in posts:
        post_id = post.get("post_id")
        print(f"Post ID: {post_id}")
    #     await db.posts.update_one(
    #         {"_id": post["_id"]},
    #         {"$set": {"analysis.scam_type": "None"}}
    #     )
    #     print(f"Updated post_id {post_id} scam_type to 'None'.")
        content = post.get("content", "")
        print(f"\nPost ID: {post_id}")
        print(f"Content: {content}\n")
        value = input("Enter value for analysis.scam_framing2 (or leave blank to skip): ").strip()
        if value:
            await db.posts.update_one(
                {"_id": post["_id"]},
                {"$set": {"analysis.scam_framing2": value}}
            )
            print(f"Updated post_id {post_id} with scam_framing2: {value}")
        else:
            print(f"Skipped post_id {post_id}")

    # print("Scam type normalization completed.")

async def list_unique_scam_framing2():
    """
    List all unique values in the analysis.scam_framing2 field from the posts collection.
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]
    unique_values = await db.posts.distinct("analysis.scam_framing2")
    print("Unique values in analysis.scam_framing2:")
    for value in unique_values:
        print(value)
    return unique_values

async def manual_update_missing_scam_framing2():
    """
    For each post missing analysis.scam_framing2, print post_id and content,
    prompt user to type in the value, and update analysis.scam_framing2 in the database.
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]
    query = {
        "$or": [
            {"analysis.scam_framing2": {"$exists": False}},
            {"analysis.scam_framing2": None},
            {"analysis.scam_framing2": "nown"}
        ]
    }
    posts = await db.posts.find(query).to_list(length=None)
    print("Posts missing analysis.scam_framing2:")

    for post in posts:
        post_id = post.get("post_id")
        print(f"Post ID: {post_id}")
        content = post.get("content", "")
        print(f"\nPost ID: {post_id}")
        print(f"Content: {content}\n")
        value = input("Enter value for analysis.scam_framing2 (or leave blank to skip): ").strip()
        if value:
            await db.posts.update_one(
                {"_id": post["_id"]},
                {"$set": {"analysis.scam_framing2": value}}
            )
            print(f"Updated post_id {post_id} with scam_framing2: {value}")
        else:
            print(f"Skipped post_id {post_id}")

    print(f"Total processed: {len(posts)}")

from openpyxl import load_workbook

def find_post_ids_not_in_excel_and_not_in_list():
    # The provided list of post_ids
    id_list = [
        901, 902, 905, 924, 928, 931, 932, 933, 934, 935, 936, 942, 943, 944, 945, 946, 947, 948,
        959, 960, 961, 962, 963, 964, 968, 972, 973, 974, 975, 976, 977, 978, 979, 980, 981, 982,
        983, 984, 985, 986, 987, 988, 989, 990, 991, 996, 997, 998, 999, 1000, 1001, 1002, 1003,
        1004, 1009, 1010, 1011, 1012, 1013, 1014, 1030, 1036, 1042, 1043, 1044, 1304, 1305, 1306,
        1310, 1313, 1314, 1315, 1316, 1317, 1321, 1322, 1323, 1342, 1346, 1349, 1360, 1361, 1370,
        1371, 1372, 1373, 1382, 1386, 1401, 1473, 1474, 1477, 1478, 1479, 1480, 1481, 1482
    ]
    id_set = set(id_list)

    # Excel file path
    file = r"C:\Users\yscha\OneDrive - matrik\USM\Y4S1\FYP\code\bend\excel\Scam_Coding_Sheet_Comparison2_Updated.xlsx"
    from openpyxl import load_workbook
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # Collect all post_ids from column A (assuming header in row 1)
    excel_post_ids = set()
    for row in range(2, ws.max_row + 1):
        post_id = ws.cell(row=row, column=1).value
        if post_id is not None:
            try:
                post_id_int = int(post_id)
                excel_post_ids.add(post_id_int)
            except Exception:
                continue

    # Check numbers from 150 to 1470 that are NOT in Excel and NOT in the list
    not_in_excel_and_not_in_list = []
    for pid in range(150, 1471):
        if pid not in id_set and pid not in excel_post_ids:
            not_in_excel_and_not_in_list.append(pid)

    print("Post IDs from 150 to 1470 NOT in the provided list and NOT in the Excel file:")
    for pid in not_in_excel_and_not_in_list:
        print(pid)
    print(f"Total: {len(not_in_excel_and_not_in_list)}")

async def update_post_dates_for_user_range():
    """
    For all posts with user_id between 11 and 15, prompt for date input
    if the date field is empty/missing, and prompt for post_url if post_url is empty/missing.
    Store the date in 'date' field in ISO format (YYYY-MM-DD) and update post_url if provided.
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Find posts with user_id between 11 and 15
    posts = await db.posts.find({"user_id": {"$gte": 11, "$lte": 15}}).to_list(length=None)

    for post in posts:
        post_id = post.get("post_id")
        post_url = post.get("post_url", "")
        post_content = post.get("content", "")
        date_value = post.get("date", "")

        update_fields = {}

        # Prompt for post_url if missing or empty
        if not post_url:
            print(f"\nPost ID: {post_id}")
            print(f"Content: {post_content}")
            print(f"Current URL: {post_url}")
            url_str = input("Enter URL for this post (or leave blank to skip): ").strip()
            if url_str:
                update_fields["post_url"] = url_str

        # Prompt for date if missing or empty
        if not date_value:
            print(f"\nPost ID: {post_id}")
            print(f"Content: {post_content}")
            print(f"Post URL: {update_fields.get('post_url', post_url)}")
            print(f"Current Date: {date_value}")
            date_str = input("Enter date for this post (dd-mm-yyyy) (or leave blank to skip): ").strip()
            if date_str:
                try:
                    # Parse and convert to ISO format (YYYY-MM-DD)
                    date_obj = datetime.strptime(date_str, "%d-%m-%Y")
                    iso_date = date_obj.date().isoformat()
                    update_fields["date"] = iso_date
                except ValueError:
                    print("Invalid date format. Skipping date update for this post.")

        # Update the document if there are any fields to update
        if update_fields:
            await db.posts.update_one(
                {"_id": post["_id"]},
                {"$set": update_fields}
            )
            print(f"Updated post_id {post_id} with: {update_fields}")
        else:
            print(f"Skipped post_id {post_id} (no updates)")

    print("Date and URL update process completed.")
    
async def print_post_counts_by_user_id():
    """
    Print the number of post documents for each user_id from 1 to 15.
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    for user_id in range(1, 16):
        count = await db.posts.count_documents({
            "user_id": user_id,
            "post_id": {"$lte": 1470}})
        print(f"user_id {user_id}: {count} posts")

async def print_comment_counts_by_user_id():
    """
    For each user_id from 1 to 15, count the total number of comments (comments.comment_content)
    by linking post_id in posts to post_id in comments.
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Build a mapping from post _id to user_id
    posts = await db.posts.find({"user_id": {"$gte": 6, "$lte": 15}}).to_list(length=None)
    postid_to_userid = {post["_id"]: post["user_id"] for post in posts}

    # Initialize counts
    user_comment_counts = {user_id: 0 for user_id in range(6, 16)}

    # Go through all comments and sum up comment_content counts by user_id
    async for comment_doc in db.comments.find({}):
        post_obj_id = comment_doc.get("post_id")
        if post_obj_id in postid_to_userid:
            user_id = postid_to_userid[post_obj_id]
            comments_list = comment_doc.get("comments", [])
            user_comment_counts[user_id] += len(comments_list)

    # Print results
    for user_id in range(6, 16):
        print(f"user_id {user_id}: {user_comment_counts[user_id]} comments")

async def print_content_and_label_counts():
    """
    Print the count of posts with non-empty content and not deleted,
    for both scam_type and scam_framing2 label queries.
    """
    client = AsyncIOMotorClient("mongodb://localhost:27017")
    db = client["scam_db2"]

    # Count for scam_type
    count_scam_type = await db.posts.count_documents({
        "deleted": {"$ne": 1},
        "content": {"$nin": [None, ""]}
    })
    print(f"Count of posts with non-empty content and not deleted (scam_type query): {count_scam_type}")

    # Count for scam_framing2
    count_scam_framing2 = await db.posts.count_documents({
        "deleted": {"$ne": 1},
        "content": {"$nin": [None, ""]}
    })
    print(f"Count of posts with non-empty content and not deleted (scam_framing2 query): {count_scam_framing2}")

if __name__ == "__main__":
    # asyncio.run(update_user_passwords())
    # asyncio.run(export_post_labelling())
    # compare_excel_files()
    # asyncio.run(update_comment_counts_for_posts())
    # update_column_r()
    # asyncio.run(update_scam_framing())
    # asyncio.run(update_sentiment_analysis_by_comment_id())
    # asyncio.run(fix_scam_framing())
    # asyncio.run(list_unique_scam_framing2())
    # asyncio.run(manual_update_missing_scam_framing2())
    # find_post_ids_not_in_excel_and_not_in_list()
    # asyncio.run(update_post_dates_for_user_range())
    # asyncio.run(print_post_counts_by_user_id())
    # asyncio.run(print_comment_counts_by_user_id())
    asyncio.run(print_content_and_label_counts())